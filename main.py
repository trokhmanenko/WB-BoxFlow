import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import io
import zipfile
from typing import Union, Optional, Dict, List
from openpyxl.worksheet.worksheet import Worksheet
from functools import reduce
import math
from datetime import datetime


def open_excel(bytes_file: bytes, header: int = 0) -> Optional[List[Dict[str, pd.DataFrame]]]:
    dataframes = []

    # Попытка прочитать Excel файл напрямую
    try:
        sheets = pd.read_excel(io.BytesIO(bytes_file), sheet_name=None, header=header)
        for sheet_name, df in sheets.items():
            dataframes.append({'name': sheet_name, 'df': df})
    except Exception as e:  # noqa
        print(f"Ошибка при чтении Excel файла напрямую: {e}")

    # Попытка прочитать Excel файлы внутри ZIP-архива
    try:
        with zipfile.ZipFile(io.BytesIO(bytes_file), 'r') as z:
            for file_name in z.namelist():
                # Проверяем, что файл является Excel файлом
                if file_name.endswith(('.xls', '.xlsx')):
                    with z.open(file_name) as f:
                        try:
                            sheets = pd.read_excel(f, sheet_name=None, header=header)
                            for sheet_name, df in sheets.items():
                                dataframes.append({'name': f"{file_name} | {sheet_name}", 'df': df})
                        except Exception as e:  # noqa
                            print(f"Ошибка при чтении Excel файла '{file_name}' из архива: {e}")
    except Exception as e:  # noqa
        print(f"Ошибка при чтении ZIP-архива: {e}")

    return dataframes if dataframes else None


def is_box_capacity(df: pd.DataFrame, *args) -> Union[bool, str]:
    # Проверяем наличие необходимых столбцов
    df.columns = [col.lower() for col in df.columns]
    required_columns = {'артикул продавца', 'размер', 'баркод', 'кратность'}
    if len(df.columns) != len(required_columns) or set(df.columns) != required_columns:
        return False
    try:
        # Пытаемся преобразовать 'Кратность' к числовому типу
        df['Кратность'] = pd.to_numeric(df['Кратность'], errors='coerce')

        for col in required_columns:
            if df[col].isnull().any():
                return f"недопустимые значения в столбце '{col}'"

        # Все проверки пройдены успешно
        return True
    except Exception as e:
        return str(e)


def is_items_to_ship(df: pd.DataFrame, *args) -> Union[bool, str]:
    df.columns = [col.lower() for col in df.columns]
    required_columns = {'баркод', 'количество'}
    if len(df.columns) != len(required_columns) or set(df.columns) != required_columns:
        return False
    try:
        # Пытаемся преобразовать 'Количество' к числовому типу
        df['Количество'] = pd.to_numeric(df['Количество'], errors='coerce')

        for col in required_columns:
            if df[col].isnull().any():
                return f"недопустимые значения в столбце '{col}'"

        # Все проверки пройдены успешно
        return True
    except Exception as e:
        return str(e)


def is_boxes_id(df: pd.DataFrame, user_data) -> Union[bool, str]:
    df.columns = [col.lower() for col in df.columns]
    required_columns = {'баркод товара', 'кол-во товаров', 'шк короба', 'срок годности'}
    if len(df.columns) != len(required_columns) or set(df.columns) != required_columns:
        return False
    if df['шк короба'].nunique() != user_data["amount_of_boxes"]:
        return "Номера коробок не совпадают"
    return True


def lcm(a, b):
    """Функция для нахождения наименьшего общего кратного двух чисел."""
    return abs(a * b) // math.gcd(a, b)


def find_lcm_of_series(series):
    """Функция для нахождения наименьшего общего кратного всех чисел в pandas Series."""
    return reduce(lcm, series)


def pack_boxes(box_capacity_df, items_to_ship_df):
    # Переименовываем столбцы
    box_capacity_df.columns = ['seller_art', 'size', 'barcode', 'multiplicity']
    items_to_ship_df.columns = ['barcode', 'quantity']

    # преобразования столбцов к типу str
    box_capacity_df['barcode'] = box_capacity_df['barcode'].astype(str)
    items_to_ship_df['barcode'] = items_to_ship_df['barcode'].astype(str)

    # Объединяем таблицы по 'barcode'
    merged_df = pd.merge(items_to_ship_df, box_capacity_df, on='barcode', how='left')
    if merged_df['multiplicity'].isna().any():
        return "Отсутствуют данные о кратности для одного или нескольких товаров"

    merged_df = merged_df.sort_values(by='multiplicity', ascending=True).reset_index(drop=True)

    box_volume = find_lcm_of_series(box_capacity_df['multiplicity'])
    merged_df['volume'] = box_volume / merged_df['multiplicity']
    merged_df['leftovers'] = merged_df['quantity']
    merged_df['box'] = 1

    # Расчет количества коробок
    merged_df['total_boxes'] = merged_df['quantity'] * merged_df['volume'] // box_volume + 1
    total_boxes = int(merged_df['total_boxes'].sum())

    # Создание DataFrame для коробок
    boxes = pd.DataFrame({'box_number': range(1, total_boxes + 1), 'volume': box_volume})

    result_rows = []

    for index, row in merged_df.iterrows():
        while row['leftovers'] > 0:
            current_box = boxes.loc[boxes['volume'] >= row['volume']].iloc[0]
            max_items_in_box = current_box['volume'] // row['volume']
            items_to_place = min(row['leftovers'], max_items_in_box)

            # Уменьшаем остатки товара и объем коробки
            row['leftovers'] -= items_to_place
            current_box_index = current_box.name
            boxes.at[current_box_index, 'volume'] -= items_to_place * row['volume']

            # Добавляем информацию в result_df
            result_row = {
                'barcode': row['barcode'],
                'quantity': items_to_place,
                'box_id': current_box_index,
                'expiration_date': '',
                'seller_art': row['seller_art'],
                'size': row['size']
            }
            result_rows.append(result_row)

            # Переход к следующей коробке, если текущая заполнена
            if boxes.at[current_box_index, 'volume'] <= 0:
                boxes.drop(current_box_index, inplace=True)

    # Создаем DataFrame из накопленных результатов
    result_df = pd.DataFrame(result_rows)

    return result_df


def count_boxes(box_capacity_df, items_to_ship_df):
    # Вызов функции pack_boxes для упаковки товаров в коробки
    result_df = pack_boxes(box_capacity_df, items_to_ship_df)

    # Если результат функции pack_boxes - строка, значит возникла ошибка
    if isinstance(result_df, str):
        print(result_df)
        return result_df

    # Подсчет количества уникальных номеров коробок в результате
    total_boxes = result_df['box_id'].nunique()
    return total_boxes


def apply_table_styles(sheet: Worksheet,
                       dataframe: pd.DataFrame,
                       start_row: int,
                       list_with_widths: Optional[List[int]] = None,
                       table_border: str = 'thin',
                       header_border: str = 'thin',
                       table_alignment: Optional[Union[str, List[str]]] = None,
                       header_alignment: Optional[str] = None,
                       wrap_text_table: bool = False,
                       wrap_text_header: bool = False,
                       cell_colors: Optional[Dict[str, List[str]]] = None,
                       range_colors: Optional[Dict[str, List[str]]] = None,
                       condition: Optional[callable] = None) -> None:
    # Set column widths
    if list_with_widths:
        for i, width in enumerate(list_with_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width

    # Set borders, alignment to table
    table_range = sheet[f"A{start_row + 1}:{chr(65 + len(dataframe.columns) - 1)}{start_row + len(dataframe)}"]
    table_border_style = Border(top=Side(border_style=table_border),
                                right=Side(border_style=table_border),
                                bottom=Side(border_style=table_border),
                                left=Side(border_style=table_border))
    if table_alignment:
        if isinstance(table_alignment, str):
            table_alignment_style = [Alignment(horizontal=table_alignment, vertical='center', wrap_text=wrap_text_table)
                                     for _ in
                                     range(dataframe.shape[1])]
        else:
            table_alignment_style = [Alignment(horizontal=t, vertical='center', wrap_text=wrap_text_table) for t in
                                     table_alignment]
    else:
        table_alignment_style = []  # Пустой список, если table_alignment не определено

    for row in table_range:
        for i, cell in enumerate(row):
            cell.border = table_border_style
            if table_alignment and i < len(table_alignment_style):
                cell.alignment = table_alignment_style[i]

    # Set borders, alignment to header
    header_range = sheet[f"A{start_row}:{chr(65 + len(dataframe.columns) - 1)}{start_row}"]
    header_border_style = Border(
        top=Side(border_style=header_border),
        right=Side(border_style=header_border),
        bottom=Side(border_style=header_border),
        left=Side(border_style=header_border)
    )
    header_alignment_style = Alignment(horizontal=header_alignment, vertical='center', wrap_text=wrap_text_header)

    for row in header_range:
        for cell in row:
            cell.border = header_border_style
            if header_alignment:
                cell.alignment = header_alignment_style

    if cell_colors and condition:
        for cell_coordinate, colors in cell_colors.items():
            cell = sheet[cell_coordinate]
            if condition(cell.value):
                cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
            else:
                cell.fill = PatternFill(start_color=colors[1], end_color=colors[1], fill_type="solid")
    elif cell_colors:
        for cell_coordinate, colors in cell_colors.items():
            cell = sheet[cell_coordinate]
            cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
    if range_colors and condition:
        for cell_range, colors in range_colors.items():
            cells = sheet[cell_range]
            for row in cells:
                for cell in row:
                    if condition(cell.value):
                        cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color=colors[1], end_color=colors[1], fill_type="solid")
    elif range_colors:
        for cell_range, colors in range_colors.items():
            cells = sheet[cell_range]
            for row in cells:
                for cell in row:
                    cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")


def generate_report(box_capacity_df, items_to_ship_df, boxes_id_df):
    packed_boxes = pack_boxes(box_capacity_df, items_to_ship_df)

    if isinstance(packed_boxes, str):
        return packed_boxes

    # Создание словаря для сопоставления box_id с "шк короба"
    box_id_to_name = dict(zip(packed_boxes['box_id'].unique(), boxes_id_df['шк короба']))

    # Обновление box_id в packed_boxes с использованием словаря
    packed_boxes['box_id'] = packed_boxes['box_id'].map(box_id_to_name)

    # # Преобразование первых двух столбцов в целочисленный тип
    # packed_boxes['barcode'] = packed_boxes['barcode'].astype(str)
    # packed_boxes['quantity'] = packed_boxes['quantity'].astype('int64')

    packed_boxes = packed_boxes.rename(columns={'barcode': 'баркод товара',
                                                'quantity': 'кол-во товаров',
                                                'box_id': 'шк короба',
                                                'expiration_date': 'срок годности',
                                                'seller_art': 'Артикул продавца',
                                                'size': 'Размер'})

    df_1_cols = ['баркод товара', 'кол-во товаров', 'шк короба', 'срок годности']
    df_2_cols = ['Артикул продавца', 'Размер'] + df_1_cols

    df_1 = packed_boxes[df_1_cols]
    df_2 = packed_boxes[df_2_cols]

    # Сохранение первого файла
    current_date = datetime.now().strftime("%Y-%m-%d")
    result_file_name1 = f'\\Раскладка коробок_{current_date}.xlsx'
    output1 = write_dfs_to_xlsx(df_1, 'Sheet1')

    # Создание и сохранение второго файла
    result_file_name2 = f'\\Инструкция для склада_{current_date}.xlsx'
    output2 = write_dfs_to_xlsx(df_2, 'Инструкция для склада')

    wb = load_workbook(filename=output2)

    apply_table_styles(
        wb['Инструкция для склада'], df_2, 1, [20] * 6,
        table_border='thin', header_border='thick', wrap_text_header=True,
        table_alignment=['left', 'right']
    )
    output2.seek(0)
    wb.save(output2)
    output2.seek(0)

    return (output1, result_file_name1), (output2, result_file_name2)


def write_dfs_to_xlsx(total: pd.DataFrame, sheet_name) -> io.BytesIO:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        total.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
    output.seek(0)
    return output
